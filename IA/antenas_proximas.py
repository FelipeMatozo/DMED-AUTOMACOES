import os
import pandas as pd
from unidecode import unidecode
from geopy.distance import geodesic
from heapq import nsmallest
from banco_utils import buscar_dados, buscar_todos_dados, ver_exist
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

class TratarPlanilha:

    def __init__(self, data_path):
        self.data_path = data_path
        self.data_df = None
        self.merged_df = None

    def carregar_planilhas(self):
        # Carregar as planilhas em DataFrames
        self.data_df = pd.read_excel(self.data_path)
        print("Planilhas carregadas com sucesso.")

    def normalizar_nomes_colunas(self, df):
        # Normalizar os nomes das colunas removendo acentos e caracteres especiais
        df.columns = [unidecode(col).strip() for col in df.columns]
        return df

    def padronizar_colunas(self, df, colunas):
        for coluna in colunas:
            df[coluna] = df[coluna].astype(str).str.strip().str.replace('.0', '', regex=False).replace('nan', '')

    def aplicar_estilos(self, path_planilha):
        # Carregar a planilha com openpyxl
        wb = load_workbook(path_planilha)
        ws = wb.active
        
        # Aplicar estilos no cabeçalho
        header_fill = PatternFill(start_color="B7DEE8", end_color="B7DEE8", fill_type="solid")
        header_font = Font(bold=True, color="000000")
        
        for cell in ws[1]:  # Primeira linha (cabeçalho)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Aplicar alinhamento às colunas numéricas
        for row in ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row):
            for cell in row:
                if isinstance(cell.value, (int, float)):  # Alinhar números à direita
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

        # Salvar a planilha com as alterações
        wb.save(path_planilha)
        print(f"Estilos aplicados e planilha salva em: {path_planilha}")
      
    def salvar_planilha(self, output_path='assets/excel/Unidades_desconectadas.xlsx'):
        

        self.processar_antenas(output_path)

       

    
    def buscar_lat_lon(self, uc):
        resultado = buscar_dados("UC_LAT_LONG", "UC", uc, ["LATITUDE", "LONGITUDE"])
        
        if resultado:
            latitude, longitude = resultado
            return float(latitude), float(longitude)
        else:
            return None

    def encontrar_antenas_mais_proximas(self, uc_lat, uc_lon, n=3):
        colunas = ["latitude_antena", "longitude_antena", "operadora", "estacao"]
        antenas = buscar_todos_dados("antenas_parana", colunas)

        distancias_antenas = []
        
        for antena_lat, antena_lon, operadora, estacao in antenas:
            distancia = geodesic((uc_lat, uc_lon), (antena_lat, antena_lon)).kilometers
            distancias_antenas.append((distancia, (operadora, estacao)))

        antenas_mais_proximas = nsmallest(n, distancias_antenas, key=lambda x: x[0])
        
        return [
            (antena[1][0], antena[1][1], f"{antena[0]:.2f} km")
            for antena in antenas_mais_proximas
        ]

    def processar_antenas(self, output_path):
        resultados = []
        print(self.data_df)
        for index, row in self.data_df.iterrows():
            numero_uc = row["UC"]
            if ver_exist("UC_LAT_LONG", "UC", numero_uc):
                print(f'A UC: {numero_uc} está presente no banco de dados!')
                latitude_uc, longitude_uc = self.buscar_lat_lon(numero_uc)

                if latitude_uc is not None and longitude_uc is not None:
                    antenas_proximas = self.encontrar_antenas_mais_proximas(latitude_uc, longitude_uc)

                    if antenas_proximas:
                        resultado_row = [numero_uc]
                        for antena in antenas_proximas:
                            resultado_row.append(f"{antena[0]} ({antena[1]})")
                            resultado_row.append(antena[2])
                        resultados.append(resultado_row)
                    else:
                        print(f"Nenhuma antena encontrada próxima à UC {numero_uc}.")
                else:
                    print(f'Não foi possível encontrar as coordenadas para a UC: {numero_uc}.')
            else:
                print(f'A UC: {numero_uc} NÃO está presente no banco de dados!')

        # Criar um DataFrame com os resultados das antenas
        colunas_resultados = ["UC", "Melhor antena", "Distância 1", "Segunda antena", "Distância 2", "Terceira antena", "Distância 3"]
        df_resultados = pd.DataFrame(resultados, columns=colunas_resultados)

        # Salvar os resultados em uma nova planilha
        resultado_path = os.path.join(os.path.dirname(output_path), "resultados_antenas.xlsx")
        df_resultados.to_excel(resultado_path, index=False)
        print(f"Resultados de antenas salvos em: {resultado_path}")

    def merge_com_planilha_antenas(self, antennas_path, output_path='assets/excel/Final_sinal_op/final_Uc_desconec.xlsx'):
        # Carregar a planilha de antenas
        antennas_df = pd.read_excel(antennas_path)

        # Realizar o merge com a planilha de antenas usando a coluna "UC"
        resultado_final = pd.merge(self.data_df, antennas_df, on='UC', how='left')  # Usar 'left' para manter todas as linhas de merged_df

        # Reorganizar as colunas na ordem desejada
        colunas_desejadas = [
            "UC","Cliente","RG","Município","MED","Operadora","Marca TM", "TM Hemera","Regional 1",
            "Melhor antena", "Distância 1", "Segunda antena", "Distância 2","Terceira antena", 
            "Distância 3"
           
        ]

        # Manter apenas as colunas desejadas que estão presentes no DataFrame
        colunas_presentes = [col for col in colunas_desejadas if col in resultado_final.columns]
        resultado_final = resultado_final[colunas_presentes]
        print("Colunas reorganizadas com sucesso.")
        # Salvar o resultado final
        resultado_final.to_excel(output_path, index=False)
        self.aplicar_estilos(output_path)
        print(f"Planilha final salva em: {output_path}")


if __name__ == "__main__":
    tratar = TratarPlanilha("assets/excel/SEM COMUNICACAO - OUTUBRO.xlsx")
    tratar.carregar_planilhas()
    tratar.salvar_planilha()
    # Caminho para a planilha de resultados de antenas
    resultados_antenas_path = "assets/excel/resultados_antenas.xlsx"
    
    # Realizar o merge com a planilha de antenas
    tratar.merge_com_planilha_antenas(resultados_antenas_path)